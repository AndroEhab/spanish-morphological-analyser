"""SUBTLEX-ESP frequency loader.

Reads all three column blocks from SUBTLEX-ESP.xlsx.
Produces word -> per-million float (sum duplicates).
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

import openpyxl


def _extract_word(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if not s or s in ("-", "---"):
        return None
    s = re.sub(r"^\d+\.\s*", "", s)
    s = s.strip()
    if not s:
        return None
    return s


def load(xlsx_path: str | Path) -> dict[str, float]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["Subtlex-Esp"]
    except KeyError:
        available = wb.sheetnames
        wb.close()
        raise ValueError(
            f"Sheet 'Subtlex-Esp' not found. Available sheets: {available}"
        )

    freq: dict[str, float] = {}

    # Read all data at once using iter_rows for performance.
    # Columns: A(word), B(raw), C(per-million) | F,G,H | K,L,M
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    for row in rows:
        for word_col_idx, pm_col_idx in [(0, 2), (5, 7), (10, 12)]:
            if word_col_idx >= len(row):
                continue
            word_cell = row[word_col_idx]
            pm_cell = row[pm_col_idx] if pm_col_idx < len(row) else None

            word = _extract_word(word_cell)
            if word is None:
                continue

            try:
                pm = float(pm_cell) if pm_cell is not None else 0.0
            except (ValueError, TypeError):
                pm = 0.0

            freq[word.lower()] = freq.get(word.lower(), 0.0) + pm

    wb.close()
    return freq
